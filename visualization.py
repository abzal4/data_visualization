import os
import pandas as pd
import numpy as np
import psycopg2
from psycopg2.extras import RealDictCursor
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule



DB = {
    "host": "localhost",
    "port": 5432,
    "dbname": "postgres",
    "user": "postgres",
    "password": "abzal"
}

def get_conn():
    conn_str = "host={host} port={port} dbname={dbname} user={user} password={password}".format(**DB)
    return psycopg2.connect(conn_str)

def run_query(sql):
    with get_conn() as conn:
        df = pd.read_sql(sql, conn)
    return df

colors = [ "#6B8F71", "#AAD2BA", "#D9FFF5", "#B9F5D8"]

def export_to_excel(dataframes_dict, filename):
    os.makedirs("exports", exist_ok=True)
    filepath = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        
        header_fill = PatternFill("solid", fgColor="FF6B8F71") 
        header_font = Font(bold=True, color="FFFFFF")       

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        df = dataframes_dict[sheet_name]
        numeric_cols = df.select_dtypes(include=["number"]).columns

        for col in numeric_cols:
            col_idx = df.columns.get_loc(col) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            data_range = f"{col_letter}2:{col_letter}{len(df)+1}"
            rule = ColorScaleRule(
                start_type="min", start_color="FFB9F5D8",
                mid_type="percentile", mid_value=50, mid_color="FFAAD2BA",
                end_type="max", end_color="FF6B8F71"
            )
            ws.conditional_formatting.add(data_range, rule)

    wb.save(filepath)

    total_sheets = len(dataframes_dict)
    total_rows = sum(len(df) for df in dataframes_dict.values())
    print("=" * 50)
    print(f"📊 Excel Report Created: {filename}")
    print(f"   📑 Sheets: {total_sheets}")
    print(f"   🔢 Total Rows: {total_rows}")
    print(f"   📂 Path: {filepath}")
    print("=" * 50)


def main():
    os.makedirs("charts", exist_ok=True)

    print("Players with Most goals in 23/24 season")
    df1 = run_query("""select c.domestic_competition_id as league,c.name as club,p.name as player, goals 
                    from (select a.player_id as id, sum(a.goals) as goals 
                          from appearances a 
                          left join games g on a.game_id = g.game_id 
                          where g.competition_type = 'domestic_league' 
                          and (a."date" between '2023-07-01' and '2024-08-10') 
                          group by a.player_id) as s
                    left join players p on s.id = p.player_id 
                    left join clubs c on p.current_club_id = c.club_id 
                    order by goals desc limit 10""")
    q1_player, q1_goals = df1['player'], df1['goals']

    # --- Bar Chart ---
    fig, ax = plt.subplots(figsize=(8,6))
    bars = ax.bar(q1_player, q1_goals, color=colors[2])
    ax.set_title("Players with Most goals in 23/24 season")
    ax.set_xlabel("Players")
    ax.set_ylabel("Goals")
    ax.set_xticklabels(q1_player, rotation=25, ha='right')
    ax.bar_label(bars, padding=3)
    plt.tight_layout()
    chart_path = os.path.join("charts", "top_goals.png")
    plt.savefig(chart_path)
    plt.close()
    print(f"✅ Bar chart saved: {chart_path}, rows: {len(df1)}. Shows Players with Most goals in 23/24 season.")

    # --- Horizontal Bar Chart (aggressiveness) ---
    df2 = run_query("""select p."name" as player, (sum(a.red_cards)*2+sum(a.yellow_cards)) as agressivness 
                       from players p 
                       inner join appearances a on p.player_id = a.player_id 
                       group by p.player_id 
                       order by agressivness desc limit 10""")
    q2_player, q2_agressivness_point = df2['player'], df2['agressivness']

    fig, ax = plt.subplots(figsize=(8,6))
    bars = ax.barh(q2_player, q2_agressivness_point, color=colors[1])
    ax.set_title("Top 10 most aggressive players")
    ax.set_xlabel("Aggressiveness points")
    ax.set_ylabel("Players")
    ax.bar_label(bars, padding=3)
    plt.tight_layout()
    chart_path = os.path.join("charts", "aggressive_players.png")
    plt.savefig(chart_path)
    plt.close()
    print(f"✅ H-bar chart saved: {chart_path}, rows: {len(df2)}. Shows Top 10 most aggressive players.")

    # --- Scatter Plot (Efficiency) ---
    df5 = run_query("""select p.name AS player, SUM(a.minutes_played) AS minutes, 
                              round(sum(a.minutes_played)/sum(a.goals),0) as min_p_goal
                       from players p
                       inner join appearances a on p.player_id = a.player_id
                       where a.minutes_played != 0 
                       and (a."date" between '2023-07-01' and '2024-08-10') 
                       and p.position = 'Attack' and p.sub_position = 'Centre-Forward'
                       group by p.player_id
                       having sum(a.goals) > 0 and SUM(a.minutes_played)>300
                       order by min_p_goal 
                       limit 50;""")

    plt.figure(figsize=(12,10))
    plt.scatter(df5['minutes'], df5['min_p_goal'], color=colors[0], s=20)
    for i, row in df5.iterrows():
        plt.text(row['minutes'], row['min_p_goal'], row['player'], fontsize=6, ha='right')
    plt.title("Player efficiency (minutes per goal)")
    plt.xlabel("Minutes played")
    plt.ylabel("Minutes per goal")
    plt.grid(True)
    plt.gca().invert_yaxis()
    chart_path = os.path.join("charts", "efficiency_scatter.png")
    plt.savefig(chart_path)
    plt.close()
    print(f"✅ Scatter plot saved: {chart_path}, rows: {len(df5)}. Shows atacking player's efficiency (minutes per goal)")

    # --- Histogram (Barca goals per match) ---
    df6 = run_query("""select g.game_id, g.date, c.name as club,  
                              case when c.club_id = g.home_club_id then g.home_club_goals
                                   when c.club_id = g.away_club_id then g.away_club_goals end as goals_scored
                       from games g
                       join clubs c on c.club_id = g.home_club_id or c.club_id = g.away_club_id
                       where c.name = 'Futbol Club Barcelona' 
                       and g."date" between '2023-07-01' and '2024-08-10'
                       order by g.date;""")
    q6_goals = df6['goals_scored']

    plt.figure(figsize=(7,5))
    plt.hist(q6_goals, bins=6, color=colors[1], alpha=0.7, edgecolor='black')
    plt.title("Barca goals per match (23/24)")
    plt.xlabel("Goals in match")
    plt.ylabel("Frequency")
    chart_path = os.path.join("charts", "barca_goals_hist.png")
    plt.savefig(chart_path)
    plt.close()
    print(f"✅ Histogram saved: {chart_path}, rows: {len(df6)}. Shows Barca goals per match (23/24)")

    # --- Pie Chart (Real Madrid goals) ---
    df7 = run_query("""select  p.name as player, sum(a.goals) as total_goals
                       from appearances a
                       join players p on a.player_id = p.player_id
                       join clubs c on p.current_club_id = c.club_id
                       where c.name = 'Real Madrid Club de Fútbol'
                       and a.date between '2023-07-01' and '2024-08-10'
                       group by p.player_id
                       having sum(a.goals)>0
                       order by total_goals desc;""")
    plt.figure(figsize=(6,6))
    plt.pie(df7['total_goals'], labels=df7['player'], autopct="%1.1f%%", startangle=140,colors=[colors[i % len(colors)] for i in range(len(df7['player']))])
    plt.title("Real Madrid goals distribution (23/24)")
    chart_path = os.path.join("charts", "real_madrid_pie.png")
    plt.savefig(chart_path)
    plt.close()
    print(f"✅ Pie chart saved: {chart_path}, rows: {len(df7)}.Shows Real Madrid goals distribution (23/24).")

    # --- Line Chart (PSG attendance) ---
    df8 = run_query("""select extract(year from g.date) as season_year, max(g.attendance) as max_attendance
                       from games g
                       join clubs c on c.club_id = g.home_club_id
                       where c.name = 'Paris Saint-Germain Football Club'
                       group by season_year
                       order by season_year;""")
    plt.figure(figsize=(7,5))
    plt.plot(df8['season_year'], df8['max_attendance'], marker='o', color='green')
    plt.title("PSG max attendance over years")
    plt.xlabel("Year")
    plt.ylabel("Max Attendance")
    plt.grid(True)
    chart_path = os.path.join("charts", "psg_attendance_line.png")
    plt.savefig(chart_path)
    plt.close()
    print(f"✅ Line chart saved: {chart_path}, rows: {len(df8)}. Shows PSG max attendance over years.")

    q9 = """WITH player_values AS (
                    SELECT   p.player_id,  p.name AS player, pv.date,  pv.market_value_in_eur
                    FROM player_valuations pv
                    JOIN players p ON pv.player_id = p.player_id
                    WHERE p.name IN ('Ousmane Dembélé','Pedri','Raphinha', 'Mohamed Salah', 'Vitinha')
                    AND pv.date BETWEEN '2019-07-01' AND '2024-08-10'
                )
                SELECT   player,  EXTRACT(YEAR FROM date) AS year,   round(AVG(market_value_in_eur),2) AS avg_value_eur
                FROM player_values
                GROUP BY player, EXTRACT(YEAR FROM date)
                ORDER BY year, player;"""
    #Time Slider
    df9= pd.DataFrame(run_query(q9))
    colors2 = {
            "Ousmane Dembélé": "#6B8F71",  # dark green
            "Pedri": "#AAD2BA",            # light green
            "Raphinha": "#D9FFF5",          # mint
            'Mohamed Salah': "#b9f5d8",
            'Vitinha': "#7DCCA6"
        }
    fig = px.bar(
        df9,
        x="player",
        y="avg_value_eur",
        color="player",
        animation_frame="year",
        title="Average Market Value of 24/25 Ballon D'or pretenders per Year",
        color_discrete_map=colors2)
    
    fig.update_layout(
        yaxis=dict(title="Average Market Value (€)"), 
        xaxis=dict(title="Player"))
    fig.show()

    
    #Export to Excel with Formatting
    dataframes = {
        "Top scorers": df1,
        "Aggressive players": df2,
        "Efficiency": df5,
        "Barca goals": df6,
        "Real Madrid goals": df7,
        "PSG attendance": df8,
        "Market values": df9
        }
    export_to_excel(dataframes, "football_report.xlsx")

    

    
if __name__ == "__main__":
    main()
